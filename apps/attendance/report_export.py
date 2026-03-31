"""
Enhanced export engine for the reporting feature.
Supports CSV, Excel (multi-sheet), and PDF (table-only, landscape).
All exports are on-demand and returned directly as HTTP responses (no file storage).
"""
import csv
import io
from django.http import HttpResponse
from django.utils.timezone import localtime
from django.utils import timezone
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Helpers ────────────────────────────────────────────────────────────────────

STATUS_COLORS = {
    'PRESENT': colors.HexColor('#166534'),
    'ABSENT':  colors.HexColor('#991b1b'),
    'LATE':    colors.HexColor('#9a3412'),
    'EXCUSED': colors.HexColor('#854d0e'),
}


def _pdf_header(elements, title, subtitle, styles):
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold', fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        alignment=1, spaceAfter=6,
    )
    sub_style = ParagraphStyle(
        'ReportSub',
        parent=styles['Normal'],
        fontName='Helvetica', fontSize=9,
        textColor=colors.HexColor('#6b7280'),
        alignment=1, spaceAfter=16,
    )
    elements.append(Paragraph("SmartAttend — Official Attendance Report", title_style))
    elements.append(Paragraph(title, sub_style))
    elements.append(Paragraph(subtitle, sub_style))
    elements.append(Spacer(1, 8))


def _apply_table_style(table, data, status_col_idx=None):
    style = TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#2563eb')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  9),
        ('ALIGN',         (0, 0), (-1, 0),  'CENTER'),
        ('TOPPADDING',    (0, 0), (-1, 0),  9),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  9),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('ALIGN',         (0, 1), (-1, -1), 'CENTER'),
        ('TOPPADDING',    (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('INNERGRID',     (0, 0), (-1, -1), 0.2, colors.HexColor('#e5e7eb')),
        ('BOX',           (0, 0), (-1, -1), 0.5, colors.HexColor('#9ca3af')),
    ])
    for i in range(1, len(data)):
        bg = colors.HexColor('#f8fafc') if i % 2 == 0 else colors.white
        style.add('BACKGROUND', (0, i), (-1, i), bg)
        if status_col_idx is not None:
            val = data[i][status_col_idx]
            tc = STATUS_COLORS.get(val, colors.black)
            style.add('TEXTCOLOR', (status_col_idx, i), (status_col_idx, i), tc)
            style.add('FONTNAME',  (status_col_idx, i), (status_col_idx, i), 'Helvetica-Bold')
    table.setStyle(style)


def _rate_color(rate):
    """Return openpyxl fill color based on attendance rate."""
    if rate >= 75:
        return PatternFill('solid', fgColor='D1FAE5')  # green
    elif rate >= 50:
        return PatternFill('solid', fgColor='FEF3C7')  # yellow
    else:
        return PatternFill('solid', fgColor='FEE2E2')  # red


# ── Report CSV ────────────────────────────────────────────────────────────────

def export_report_csv(rows, section_label, filename='attendance_report.csv'):
    """
    rows: list of dicts from get_section_student_rows()
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow([
        'Student Name', 'Roll No', 'Total Classes Offered',
        'Attended', 'Missed', 'Attendance Rate (%)', 'Cohort Avg (%)', 'Delta vs Cohort (%)'
    ])

    for r in rows:
        writer.writerow([
            r['name'], r['roll_no'], r['total_offered'],
            r['attended'], r['missed'],
            r['rate'], r.get('cohort_avg', ''), r.get('delta', ''),
        ])
    return response


# ── Report Excel (Multi-Sheet) ─────────────────────────────────────────────────

def export_report_excel(rows, section_label, detail_records=None, filename='attendance_report.xlsx'):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws1 = wb.active
    ws1.title = 'Summary'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='2563EB')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )

    # Title row
    ws1.merge_cells('A1:H1')
    ws1['A1'] = f'Attendance Summary — {section_label}'
    ws1['A1'].font = Font(bold=True, size=13, color='1E3A8A')
    ws1['A1'].alignment = Alignment(horizontal='center')

    ws1.merge_cells('A2:H2')
    ws1['A2'] = f'Generated: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1['A2'].font = Font(italic=True, size=9, color='6B7280')
    ws1['A2'].alignment = Alignment(horizontal='center')

    headers = ['#', 'Student Name', 'Roll No', 'Total Offered', 'Attended', 'Missed', 'Rate (%)', 'vs Cohort (%)']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    for i, r in enumerate(rows, 1):
        row_data = [
            i, r['name'], r['roll_no'], r['total_offered'],
            r['attended'], r['missed'], r['rate'], r.get('delta', 0)
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i + 4, column=col, value=val)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin
            if col == 7:  # Rate column
                cell.fill = _rate_color(r['rate'])
                cell.font = Font(bold=True)

    # Column widths
    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 18
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws1.column_dimensions[col].width = 16

    # ── Sheet 2: Detailed Records ──
    if detail_records:
        ws2 = wb.create_sheet('Detailed Records')
        det_headers = ['Student Name', 'Roll No', 'Date', 'Time', 'Course', 'Section', 'Status', 'Mode']
        for col, h in enumerate(det_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin

        for i, rec in enumerate(detail_records, 2):
            start_time = localtime(rec.session.started_at)
            status = rec.status
            status_fills = {
                'PRESENT': PatternFill('solid', fgColor='D1FAE5'),
                'ABSENT':  PatternFill('solid', fgColor='FEE2E2'),
                'LATE':    PatternFill('solid', fgColor='FEF3C7'),
                'EXCUSED': PatternFill('solid', fgColor='DBEAFE'),
            }
            row_vals = [
                rec.student.full_name,
                rec.student.university_roll_number or rec.student.student_id,
                start_time.strftime('%Y-%m-%d'),
                start_time.strftime('%H:%M'),
                rec.session.section.course_code,
                rec.session.section.section_identifier,
                status,
                rec.session.mode,
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws2.cell(row=i, column=col, value=val)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin
                if col == 7:
                    cell.fill = status_fills.get(status, PatternFill())
                    cell.font = Font(bold=True)

        for col_letter, width in zip('ABCDEFGH', [28, 18, 12, 8, 12, 10, 10, 12]):
            ws2.column_dimensions[col_letter].width = width

    wb.save(response)
    return response


# ── Report PDF ────────────────────────────────────────────────────────────────

def export_report_pdf(rows, section_label, detail_records=None, filename='attendance_report.pdf'):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(
        response, pagesize=landscape(letter),
        rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28
    )
    elements = []
    styles = getSampleStyleSheet()

    # Header
    gen_time = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    _pdf_header(
        elements,
        title=f'Section: {section_label}',
        subtitle=f'Generated: {gen_time}  |  Total Students: {len(rows)}',
        styles=styles,
    )

    # Summary Table
    headers = ['#', 'Student Name', 'Roll No', 'Total Offered', 'Attended', 'Missed', 'Rate %', 'vs Cohort %']
    data = [headers]
    for i, r in enumerate(rows, 1):
        data.append([
            str(i),
            r['name'],
            r['roll_no'],
            str(r['total_offered']),
            str(r['attended']),
            str(r['missed']),
            f"{r['rate']}%",
            f"{r.get('delta', 0):+.1f}%",
        ])

    col_widths = [0.4*inch, 2.5*inch, 1.5*inch, 1.1*inch, 1.0*inch, 0.9*inch, 0.8*inch, 1.0*inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#2563eb')),
        ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',      (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, 0),  9),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING',    (0, 0), (-1, 0),  9),
        ('BOTTOMPADDING', (0, 0), (-1, 0),  9),
        ('FONTNAME',      (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('TOPPADDING',    (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('INNERGRID',     (0, 0), (-1, -1), 0.2, colors.HexColor('#e5e7eb')),
        ('BOX',           (0, 0), (-1, -1), 0.5, colors.HexColor('#9ca3af')),
    ])

    for i, r in enumerate(rows, 1):
        bg = colors.HexColor('#f8fafc') if i % 2 == 0 else colors.white
        style.add('BACKGROUND', (0, i), (-1, i), bg)
        rate = r['rate']
        if rate >= 75:
            rate_color = colors.HexColor('#166534')
        elif rate >= 50:
            rate_color = colors.HexColor('#854d0e')
        else:
            rate_color = colors.HexColor('#991b1b')
        style.add('TEXTCOLOR', (6, i), (6, i), rate_color)
        style.add('FONTNAME',  (6, i), (6, i), 'Helvetica-Bold')

        delta = r.get('delta', 0)
        delta_color = colors.HexColor('#166534') if delta >= 0 else colors.HexColor('#991b1b')
        style.add('TEXTCOLOR', (7, i), (7, i), delta_color)
        style.add('FONTNAME',  (7, i), (7, i), 'Helvetica-Bold')

    table.setStyle(style)
    elements.append(table)

    # Detailed Records Section (if provided)
    if detail_records:
        elements.append(Spacer(1, 20))
        det_title_style = ParagraphStyle(
            'DetTitle', parent=styles['Heading2'],
            fontName='Helvetica-Bold', fontSize=11,
            textColor=colors.HexColor('#1e3a8a'), spaceAfter=8,
        )
        elements.append(Paragraph("Detailed Attendance Records", det_title_style))

        det_headers = ['Date', 'Time', 'Student', 'Roll No', 'Course', 'Section', 'Status']
        det_data = [det_headers]
        for rec in detail_records:
            start_time = localtime(rec.session.started_at)
            det_data.append([
                start_time.strftime('%Y-%m-%d'),
                start_time.strftime('%H:%M'),
                rec.student.full_name,
                rec.student.university_roll_number or rec.student.student_id,
                rec.session.section.course_code,
                rec.session.section.section_identifier,
                rec.status,
            ])

        det_col_widths = [1.0*inch, 0.8*inch, 2.2*inch, 1.4*inch, 1.1*inch, 0.9*inch, 0.9*inch]
        det_table = Table(det_data, colWidths=det_col_widths, repeatRows=1)
        _apply_table_style(det_table, det_data, status_col_idx=6)
        elements.append(det_table)

    doc.build(elements)
    return response
